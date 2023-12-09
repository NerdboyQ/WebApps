#### The following code was developed as a Scraper for the Maryland Estates Records
#### Website. This code will open a browser in Firefox and collect records for a 
#### specified date range. This collected data will be exported to an excel file.
#### NOTE: Each time the scraper runs it will overwrite the previous excel file.

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import NoSuchAttributeException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import time
import asyncio, sys, os, requests, datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from os import system

def main(site, driver, wb, links, startDate, endDate):
    c_delay = 5
    driver.implicitly_wait(10)
    ws = wb.create_sheet("Main")
    ws.cell(1,1,"County")
    ws.cell(1,2,"Estate Number")
    ws.cell(1,3,"Filing Date")
    ws.cell(1,4,"Date of Death")
    ws.cell(1,5,"Type")
    ws.cell(1,6,"Status")
    ws.cell(1,7,"Name")
    startDate_field = driver.find_element("xpath", "//input[@id='DateOfFilingFrom']")
    endedDate_field = driver.find_element("xpath", "//input[@id='DateOfFilingTo']")
    
    startDate_field.send_keys(startDate)
    endedDate_field.send_keys(endDate)


    driver.find_element("id", "cmdSearch").click()
    time.sleep(c_delay)
    page_count = driver.find_element("xpath", "//td[@class='status']").text
    page_count = page_count.split("of ")
    page_count = page_count[1].split(" (")
    page_count = int(page_count[0])
    xl_row = 2
    print("number of pages = "+str(page_count))
    if page_count > 10:
        p_link = 1
        test_mode = 0
        max_10Umulti = (page_count/10)*10
        print("begin max 10 page = "+str(max_10Umulti))
        for n in range (2,page_count+2):
            print("n: "+str(n))
            ####test_block
            if test_mode == 1:
                p_link = 2
                maxi = page_count/10 - 1
                print("begin maxi "+str(maxi))
                testFfwd(driver, c_delay,maxi)
                test_mode = 0
                print("begin n: "+str(n))
                print("begin page_count: "+str(page_count))
            ####test_block end
            if n < 12:
                rows = driver.find_elements("xpath", "//table[@id='dgSearchResults']/tbody/tr")
                for row in range(1,len(rows)-1):
                    links.append(rows[row].find_element("xpath", ".//a").get_attribute("href"))
                    cols = rows[row].find_elements("xpath", ".//td")
                    for col in range(0,len(cols)):
                        ws.cell(xl_row,col+1,cols[col].text.encode("utf-8"))
                    xl_row+=1
                pager_row = driver.find_element("xpath", "//tr[@class='grid-pager']")
                if p_link<10:
                    print("clicking on page "+pager_row.find_element("xpath", ".//a["+str(p_link)+"]").text)
                    pager_row.find_element("xpath", ".//a["+str(p_link)+"]").click()

                    time.sleep(c_delay/2)
                    p_link+=1
                    #print "---1 page link: "+str(p_link)
                else:
                    #print "---1 breaking at page link: "+str(p_link)
                    print("---1 clicking on page "+pager_row.find_element("xpath", ".//a["+str(p_link)+"]").text)
                    pager_row = driver.find_element("xpath", "//tr[@class='grid-pager']")
                    pages = pager_row.find_elements("xpath", ".//a")
                    pages[len(pages)-1].click()
                    time.sleep(c_delay/2)
                    p_link = 2
            elif n > 11 and n < max_10Umulti+2:
                rows = driver.find_elements("xpath", "//table[@id='dgSearchResults']/tbody/tr")
                #print "rows count = " + str(len(rows))
                for row in range(1,len(rows)-1):
                    #print "trying for loop"
                    try: 
                        links.append(rows[row].find_element("xpath", ".//a").get_attribute("href"))
                    except Exception:
                        links.append("N/A")
                        print("link error")
                    #print rows[row].find_element("xpath", ".//a").text+", link : "+rows[row].find_element("xpath", ".//a").get_attribute("href")
                    cols = rows[row].find_elements("xpath", ".//td")
                    for col in range(0,len(cols)):
                        ws.cell(xl_row,col+1,cols[col].text.encode("utf-8"))
                    xl_row+=1
                pager_row = driver.find_element("xpath", "//tr[@class='grid-pager']")
                if p_link<11:
                    print("---2 clicking on page "+pager_row.find_element("xpath", ".//a["+str(p_link)+"]").text)
                    pager_row.find_element("xpath", ".//a["+str(p_link)+"]").click()

                    time.sleep(c_delay/2)
                    p_link+=1
                    #print "---2 page link: "+str(p_link)
                else:
                    #print "---2 breaking at page link: "+str(p_link)
                    print("---2 clicking on page "+pager_row.find_element("xpath", ".//a["+str(p_link)+"]").text)
                    pager_row = driver.find_element("xpath", "//tr[@class='grid-pager']")
                    pages = pager_row.find_elements("xpath", ".//a")
                    pages[len(pages)-1].click()
                    time.sleep(c_delay/2)
                    p_link = 2
            else:
                rows = driver.find_elements("xpath", "//table[@id='dgSearchResults']/tbody/tr")
                for row in range(1,len(rows)-1):
                    try: 
                        links.append(rows[row].find_element("xpath", ".//a").get_attribute("href"))
                    except Exception:
                        links.append("N/A")
                        print("link error")
                    #print rows[row].find_element("xpath", ".//a").text+", link : "+rows[row].find_element("xpath", ".//a").get_attribute("href")
                    cols = rows[row].find_elements("xpath", ".//td")
                    for col in range(0,len(cols)):
                        ws.cell(xl_row,col+1,cols[col].text.encode("utf-8"))
                    xl_row+=1
                pager_row = driver.find_element("xpath", "//tr[@class='grid-pager']")
                if n<page_count+1:
                    print("---3 clicking on page "+pager_row.find_element_by_partial_link_text(str(n)).text)
                    pager_row.find_element_by_partial_link_text(str(n)).click()

                    time.sleep(c_delay)
                else:
                    print("---3 finished: "+str(n))
    else:
        p_link = 1
        test_mode = 0
        max_10Umulti = (page_count/10)*10
        print("begin max 10 page = "+str(max_10Umulti))
        for n in range (2,page_count+2):
            print("n: "+str(n))
            ####test_block
            if test_mode == 1:
                p_link = 2
                maxi = page_count/10 - 1
                print("begin maxi "+str(maxi))
                testFfwd(driver, c_delay,maxi)
                test_mode = 0
                print("begin n: "+str(n))
                print("begin page_count: "+str(page_count))
            ####test_block end
            if n < 12:
                rows = driver.find_elements("xpath", "//table[@id='dgSearchResults']/tbody/tr")
                for row in range(1,len(rows)-1):
                    links.append(rows[row].find_element("xpath", ".//a").get_attribute("href"))
                    cols = rows[row].find_elements("xpath", ".//td")
                    for col in range(0,len(cols)):
                        ws.cell(xl_row,col+1,cols[col].text.encode("utf-8"))
                    xl_row+=1
                pager_row = driver.find_element("xpath", "//tr[@class='grid-pager']")
                if p_link<page_count:
                    print("clicking on page "+pager_row.find_element("xpath", ".//a["+str(p_link)+"]").text)
                    pager_row.find_element("xpath", ".//a["+str(p_link)+"]").click()

                    time.sleep(c_delay/2)
                    p_link+=1
                    #print "---1 page link: "+str(p_link)
                else:
                    print("---1 breaking at page link: "+str(p_link))
                    break

    print("number of links: "+str(len(links)))
    return links    
    #f.write(agency_name.encode("utf-8")+"\n")

def getDetailsFromLinks(link, driver, wb, ws_master, master_row):
    c_delay = 5
    print(link)
    driver.get(link)
    driver.implicitly_wait(c_delay)
    ws_name =  driver.find_element("id", "lblEstateNumber").text
    county = driver.find_element("xpath", "//*[@class='search-header-container RECORDHEADER']/tbody/tr/td").text
    county = county.split("(")
    county = county[1].replace(")","")
    county = county.replace(" County","")
    estateNumber = ws_name
    estateType = driver.find_element("id", "lblType").text
    estateStatus = driver.find_element("id", "lblStatus").text
    estateDateOpened = driver.find_element("id", "lblDateOpened").text
    estateDateClosed = driver.find_element("id", "lblDateClosed").text
    estateReference = driver.find_element("id", "lblReference").text
    estateName = driver.find_element("id", "lblName").text
    estateDateOfDeath = driver.find_element("id", "lblDateOfDeath").text
    estateDateOfFiling = driver.find_element("id", "lblDateOfFiling").text
    estateWill = driver.find_element("id", "lblWill").text
    estateDateOfWill = driver.find_element("id", "lblDateOfWill").text
    estateDateOfProbate = driver.find_element("id", "lblDateOfProbate").text
    estateAliases = driver.find_element("id", "lblAliases").text
    estatePersonalReps = driver.find_element("id", "lblPersonalReps").text
    estateAttorney = driver.find_element("id", "lblAttorney").text
    ws = wb.create_sheet(ws_name)
    print(ws_name)
    ws.cell(1,1,"Estate Number:")
    ws.cell(1,2,estateNumber)
    ws.cell(1,4,"Type:")
    ws.cell(1,5,estateType)
    ws.cell(2,1,"Status:")
    ws.cell(2,2,estateStatus)
    ws.cell(2,4,"Date Opened:")
    ws.cell(2,5,estateDateOpened)
    ws.cell(3,1,"Date Closed:")
    ws.cell(3,2,estateDateClosed)
    ws.cell(3,4,"Reference:")
    ws.cell(3,5,estateReference)
    ws.cell(4,1,"Decendent Name:")
    ws.cell(4,2,estateName)
    ws.cell(5,1,"Date of Death")
    ws.cell(5,2,estateDateOfDeath)
    ws.cell(5,4,"Date of Filing")
    ws.cell(5,5,estateDateOfFiling)
    ws.cell(6,1,"Will:")
    ws.cell(6,2,estateWill)
    ws.cell(6,4,"Date of Will:")
    ws.cell(6,5,estateDateOfWill)
    ws.cell(7,1,"Date of Probate:")
    ws.cell(7,2,estateDateOfProbate)
    ws.cell(8,1,"Aliases:")
    ws.cell(8,2,estateAliases)
    ws.cell(9,1,"Personal Reps")
    ws.cell(9,2,estatePersonalReps)
    ws.cell(10,1,"Attorney:")
    ws.cell(10,2,estateAttorney)


    ws_master.cell(master_row,1,county)
    if estatePersonalReps != "":
        PersonalReps = driver.find_element("id", "lblPersonalReps").text
        reps = PersonalReps.split(" [")
        if len(reps) > 1:
            rep = reps[0]
            repAddr = reps[1]
        else:
            rep = reps[0]
            repAddr = "N/A"
    else:
        rep = ""
        repAddr = ""
    if estateAttorney != "":
        Attorney = driver.find_element("id", "lblAttorney").text
        atts = Attorney.split(" [")
        if len(atts) > 1:
            att = atts[0]
            attAddr = atts[1][:-1]
        else:
            att = atts[0]
            attAddr = "N/A"
    else:
        att = ""
        attAddr = ""
    master_arr_fields = [county, estateType, estateName, rep, repAddr, att, attAddr, estateNumber]
    for col in range(0, len(master_arr_fields)):
        ws_master.cell(master_row,col+1,master_arr_fields[col])

    ws.cell(11,1,"Docket History")
    xl_row = 12
    rows = driver.find_elements("xpath", "//*[@class='docket-history-data']/tbody/tr") 

    for row in rows:
        if xl_row == 12:
            cols = row.find_elements("xpath", ".//th")
            #print len(cols)
            for col in range(1,len(cols)-1):
                ws.cell(xl_row,col,cols[col].text.encode("utf-8"))
        else:
            cols = row.find_elements("xpath", ".//td")
            #print len(cols)
            for col in range(1,len(cols)-1):
                ws.cell(xl_row,col,cols[col].text.encode("utf-8"))
        xl_row+=1
    wb.save('MarylandEstates_Report.xlsx')

def testFfwd(driver, c_delay, maxi):
        for n in range (0,maxi):
            pager_row = driver.find_element("xpath", "//tr[@class='grid-pager']")
            pages = pager_row.find_elements("xpath", ".//a")
            pages[len(pages)-1].click()
            time.sleep(c_delay/3)

def getArgs():
    args = sys.argv
    arr = []
    if len(args) == 3:
        arr.append(args[1])
        arr.append(args[2])
    elif len(args) == 2:
        arr.append(args[1])
        endedDate = str(datetime.date.today()).split("-")
        endedDate = endedDate[1]+"/"+endedDate[2]+"/"+endedDate[0]
        arr.append(endedDate)
    elif len(args) == 1:
        endDate = str(datetime.date.today()).split("-")
        startDate = str(datetime.date.today() - datetime.timedelta(1)).split("-")
        endDate = str(endDate[1])+"/"+str(endDate[2])+"/"+str(endDate[0])
        startDate = str(startDate[1])+"/"+str(startDate[2])+"/"+str(startDate[0])
        
        arr.append(startDate)
        arr.append(endDate)
    print("search range is from "+arr[0]+"-"+arr[1])
    return arr

async def run_scraper(startDate, endDate):
    run_date = datetime.datetime.now()
    start_time = run_date.time()
    print("started run at: "+str(start_time))
    global scraperResults
    scraperResults = {
        "run_date": run_date.strftime("%m/%d/%Y_%H:%M:%S"),
        "rec_cnt" : 0
    }

    site = "https://registers.maryland.gov/rownetweb/estates/frmestatesearch2.aspx"

    wb = Workbook()
    links = []
    driver = webdriver.Firefox()
    driver.get(site)
    try:
        links = main(site, driver, wb, links, startDate, endDate)
        scraperResults['rec_cnt'] = len(links)
    except Exception as e:
        print(f" ERROR: {e}")
        return scraperResults
    
    #links = ["https://registers.maryland.gov/rownetweb/estates/frmDocketImages.aspx?src=row&RecordId=502451354"]
    print("outter - number of links: "+str(len(links)))
    ws_master = wb.create_sheet("master")
    master_arr_headers = ["County","Type","Name","Personal Rep", "PR Address", "Attorney", "Attorney Address", "Number", "Marketing"]
    for header in range(0, len(master_arr_headers)):
        ws_master.cell(1,header+1,master_arr_headers[header])
    master_row = 2
    refresh_counter = 0
    for link in links:
        if refresh_counter < 10:
            refresh_counter+=1
        else:
            driver.close()
            driver = webdriver.Firefox()
            refresh_counter = 0
        getDetailsFromLinks(link, driver, wb, ws_master, master_row)
        master_row+=1
    try:
        driver.close()
    except Exception:
        print("already closed browser")
    
    
    
    fname = run_date.strftime("%d%b%Y_%Hh%Mm")
    wb.save(f"{fname}MarylandEstates_Report.xlsx")
    end_time = datetime.datetime.now().time()
    print("started run at: "+str(start_time)+"\t"+"finished run at: "+str(end_time))

    return scraperResults

scraperResults = {}
# Test
# endDate = datetime.datetime.today().strftime("%m/%d/%Y")
# startDate = (datetime.date.today() - datetime.timedelta(1)).strftime("%m/%d/%Y")
# run_scraper(startDate, endDate)